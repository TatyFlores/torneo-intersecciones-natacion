import pandas as pd
from openpyxl.styles import Alignment
from funciones import abrir_archivo, separar_secciones, separar_sexo, cantidad_series, escribir_csv, mezclar

data = abrir_archivo()

# Separar en natación 1 y 2
natacion1, natacion2 = separar_secciones(data)
fem_n1, masc_n1 = separar_sexo(natacion1)
fem_n2, masc_n2 = separar_sexo(natacion2)
series_f_1 = cantidad_series(fem_n1)
series_f_2 = cantidad_series(fem_n2)
series_m_1 = cantidad_series(masc_n1)
series_m_2 = cantidad_series(masc_n2)

print(fem_n1)

f_1 = []
f_2 = []
m_1 = []
m_2 = []

f_1 = mezclar(series_f_1, fem_n1)
f_2 = mezclar(series_f_2, fem_n2)
# m_1 = mezclar(series_m_1, masc_n1)
# m_2 = mezclar(series_m_2, masc_n2)

archivos = ["femenino_1.csv", "femenino_2.csv"]#, "masculino_1.csv", "masculino_2.csv"]

escribir_csv(archivos[0], f_1)
escribir_csv(archivos[1], f_2)
# escribir_csv(archivos[2], m_1)
# escribir_csv(archivos[3], m_2)

with pd.ExcelWriter("registro.xlsx", engine="openpyxl") as writer:
    for f in archivos:
        df = pd.read_csv(f, sep=",", encoding="utf-8") 
        nombre_hoja = f[:-4] 
        df.to_excel(writer, sheet_name=nombre_hoja, index=False)

        ws = writer.book[nombre_hoja]
        col = 1                 # columna A
        start_row = 2           # datos empiezan en la fila 2 (fila 1 = encabezados)
        end_row = 1 + len(df)

        r = start_row
        while r <= end_row:
            val = ws.cell(row=r, column=col).value
            # extiende mientras el siguiente sea igual
            r2 = r + 1
            while r2 <= end_row and ws.cell(row=r2, column=col).value == val:
                r2 += 1
            # si hay al menos 2 filas iguales consecutivas, merge A{r}:A{r2-1}
            if r2 - 1 > r:
                ws.merge_cells(start_row=r, start_column=col, end_row=r2-1, end_column=col)
                c = ws.cell(row=r, column=col)
                c.alignment = Alignment(horizontal="center", vertical="center")
            r = r2



